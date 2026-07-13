# Audit Agent API: upload documents, upload questions, process, run audit, fetch results, export Excel.

from __future__ import annotations

import asyncio
import io
import os
import re
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import APIRouter, Header, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse

from agents.audit.agents.audit_agent import conduct_audit
from agents.audit.agents.verifier import AuditResultVerifier
from agents.audit.processors.document_extractor import (
    SUPPORTED_EXTENSIONS,
    extract_document as extract_document_file,
)
from agents.audit.processors.embedder import clear_collection, embed_pages
from agents.audit.processors.questions_parser import parse_questions_csv
from agents.audit.processors.vision_extractor import extract_vision_pages
from agents.audit.schemas import (
    AuditResponse,
    AuditResultItem,
    AuditSummary,
    ExportRequest,
    QuestionValidationResponse,
)
from core.config import settings
from core.logger import get_logger
from core.session_manager import session_manager

router = APIRouter()
logger = get_logger("audit.router")
verifier = AuditResultVerifier()

UPLOAD_PATH = Path(settings.UPLOADS_PATH)
UPLOAD_PATH.mkdir(parents=True, exist_ok=True)

SESSION_TTL = 7200  # 2 hours

# Redis field names (session_manager stores as session:{session_id}:{field})
_FIELD_FILES = "audit:dynamic_files"
_FIELD_QUESTIONS = "audit:dynamic_questions"

# Session IDs from the client end up as directory names — restrict to a safe
# charset so "../" can never escape the uploads directory.
_SESSION_ID_RE = re.compile(r"^[A-Za-z0-9_-]{8,64}$")

_MAX_UPLOAD_BYTES = settings.MAX_UPLOAD_MB * 1024 * 1024


def _require_session(session_id: str) -> str:
    if not _SESSION_ID_RE.fullmatch(session_id):
        raise HTTPException(status_code=400, detail="Invalid X-Session-ID format.")
    return session_id


def _safe_filename(filename: str) -> str:
    """Strip any path components from a client-supplied filename."""
    return Path(filename).name


def _results_field(category: str) -> str:
    return f"audit:results:{category}"


# ---------------------------------------------------------------------------
# POST /upload-documents  (dynamic flow)
# ---------------------------------------------------------------------------

@router.post("/upload-documents")
async def upload_documents(
    files: list[UploadFile] = File(...),
    x_session_id: str = Header(..., alias="X-Session-ID"),
):
    """Accept multiple document uploads (PDF/PPTX/DOCX/XLSX/CSV), save to session dir."""
    session_id = _require_session(x_session_id)
    session_dir = UPLOAD_PATH / session_id
    session_dir.mkdir(parents=True, exist_ok=True)

    uploaded: list[dict] = []

    for file in files:
        filename = _safe_filename(file.filename or "")
        ext = Path(filename).suffix.lower()

        if ext not in SUPPORTED_EXTENSIONS:
            label = ext.lstrip(".").upper() if ext else "unknown"
            raise HTTPException(
                status_code=422,
                detail=(
                    f"File type '{label}' is not supported. "
                    "Supported types: PDF, PPTX, PPT, DOCX, DOC, XLSX, CSV"
                ),
            )

        contents = await file.read()
        if len(contents) > _MAX_UPLOAD_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"'{filename}' is too large — the limit is {settings.MAX_UPLOAD_MB} MB.",
            )
        save_path = session_dir / filename
        save_path.write_bytes(contents)
        uploaded.append({"filename": filename, "type": ext.lstrip(".").upper()})
        logger.info(f"doc_uploaded session_id={session_id} filename={filename} size={len(contents)}")

    # Merge with any previously uploaded files tracked in Redis
    existing: list[dict] = (await session_manager.get(session_id, _FIELD_FILES)) or []
    existing_names = {f["filename"] for f in existing}
    for u in uploaded:
        if u["filename"] not in existing_names:
            existing.append(u)

    await session_manager.set(session_id, _FIELD_FILES, existing, ttl=SESSION_TTL)

    return {"session_id": session_id, "files": existing}


# ---------------------------------------------------------------------------
# POST /upload-questions  (dynamic flow)
# ---------------------------------------------------------------------------

@router.post("/upload-questions", response_model=QuestionValidationResponse)
async def upload_questions(
    file: UploadFile = File(...),
    x_session_id: str = Header(..., alias="X-Session-ID"),
):
    """Validate a CSV of audit questions. On success, cache parsed questions in Redis."""
    session_id = _require_session(x_session_id)

    if not file.filename or not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=422, detail="Only CSV files are accepted for questions.")

    contents = await file.read()
    if len(contents) > _MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"The questions file is too large — the limit is {settings.MAX_UPLOAD_MB} MB.",
        )

    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as tmp:
        tmp.write(contents)
        tmp_path = tmp.name

    try:
        questions = parse_questions_csv(tmp_path)
    except ValueError as exc:
        os.unlink(tmp_path)
        return QuestionValidationResponse(
            valid=False,
            question_count=0,
            preview=[],
            error=str(exc),
        )
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

    # Save the file and cache parsed questions
    session_dir = UPLOAD_PATH / session_id
    session_dir.mkdir(parents=True, exist_ok=True)
    (session_dir / "questions.csv").write_bytes(contents)

    await session_manager.set(session_id, _FIELD_QUESTIONS, questions, ttl=SESSION_TTL)

    logger.info(f"questions_uploaded session_id={session_id} count={len(questions)}")

    return QuestionValidationResponse(
        valid=True,
        question_count=len(questions),
        preview=[q["rule"] for q in questions[:3]],
        error=None,
    )


# ---------------------------------------------------------------------------
# POST /process-dynamic  (dynamic flow)
# ---------------------------------------------------------------------------

@router.post("/process-dynamic")
async def process_dynamic(
    x_session_id: str = Header(..., alias="X-Session-ID"),
):
    """Extract + embed all documents uploaded via /upload-documents for this session."""
    session_id = x_session_id

    file_records: list[dict] = await session_manager.get(session_id, _FIELD_FILES)
    if not file_records:
        raise HTTPException(
            status_code=400,
            detail="No documents uploaded. Call /upload-documents first.",
        )

    session_dir = UPLOAD_PATH / session_id

    file_summaries: list[dict] = []
    all_pages: list = []

    for record in file_records:
        filename = record["filename"]
        file_path = session_dir / filename

        if not file_path.exists():
            logger.warning(f"dynamic_file_missing filename={filename} session_id={session_id}")
            continue

        try:
            pages = await asyncio.to_thread(extract_document_file, str(file_path), filename)

            pages = await extract_vision_pages(pages)
            vision_count = sum(1 for p in pages if not p.needs_vision and p.image_bytes)

            all_pages.extend(pages)

            file_summaries.append(
                {
                    "filename": filename,
                    "type": record["type"],
                    "pages": len(pages),
                    "vision_used": vision_count,
                }
            )
            logger.info(f"dynamic_file_processed filename={filename} pages={len(pages)}")

        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc))
        except Exception as exc:
            logger.error(f"dynamic_file_error filename={filename} error={exc}")
            raise HTTPException(
                status_code=500,
                detail=f"Failed to process '{filename}': {exc}",
            )

    if not all_pages:
        raise HTTPException(
            status_code=400,
            detail="No content could be extracted from the uploaded files.",
        )

    try:
        chunks_stored = await asyncio.to_thread(embed_pages, session_id, all_pages)
    except Exception as exc:
        # Embedding depends on VoyageAI — surface the real cause (bad/missing
        # API key, rate limit, network) instead of an opaque 500.
        logger.error(f"dynamic_embed_error session_id={session_id} error={exc}")
        raise HTTPException(
            status_code=502,
            detail=f"Embedding failed while indexing the documents: {exc}",
        ) from exc

    logger.info(f"dynamic_process_complete session_id={session_id} files={len(file_summaries)} pages={len(all_pages)} chunks={chunks_stored}")

    return {
        "session_id": session_id,
        "files": file_summaries,
        "total_pages": len(all_pages),
        "chunks_stored": chunks_stored,
    }


# ---------------------------------------------------------------------------
# POST /audit-dynamic  (dynamic flow)
# ---------------------------------------------------------------------------

@router.post("/audit-dynamic", response_model=AuditResponse)
async def run_audit_dynamic(
    x_session_id: str = Header(..., alias="X-Session-ID"),
):
    """Run audit using the questions cached from /upload-questions."""
    session_id = x_session_id

    questions: list[dict] = await session_manager.get(session_id, _FIELD_QUESTIONS)
    if not questions:
        raise HTTPException(
            status_code=400,
            detail="No questions found for this session. Upload a questions CSV via /upload-questions first.",
        )

    category = "Custom Audit"

    logger.info(f"dynamic_audit_start session_id={session_id} rules={len(questions)}")

    raw_results = await conduct_audit(session_id, questions)

    await session_manager.set(session_id, _results_field(category), raw_results, ttl=SESSION_TTL)

    result_items = [AuditResultItem(**r) for r in raw_results]
    summary_dict = verifier.get_result_summary(raw_results)
    summary = AuditSummary(**summary_dict)

    logger.info(f"dynamic_audit_complete session_id={session_id}")

    return AuditResponse(
        session_id=session_id,
        category=category,
        results=result_items,
        summary=summary,
    )


# ---------------------------------------------------------------------------
# GET /results/{session_id}/{category}
# ---------------------------------------------------------------------------

@router.get("/results/{session_id}/{category}", response_model=AuditResponse)
async def get_results(session_id: str, category: str):
    """Retrieve previously computed audit results."""
    raw_results: list[dict] = await session_manager.get(session_id, _results_field(category))
    if raw_results is None:
        raise HTTPException(status_code=404, detail="No results found. Run /audit-dynamic first.")

    result_items = [AuditResultItem(**r) for r in raw_results]
    summary_dict = verifier.get_result_summary(raw_results)
    summary = AuditSummary(**summary_dict)

    return AuditResponse(
        session_id=session_id,
        category=category,
        results=result_items,
        summary=summary,
    )


# ---------------------------------------------------------------------------
# POST /export
# ---------------------------------------------------------------------------

@router.post("/export")
async def export_results(req: ExportRequest):
    """Generate and stream an Excel report for the given session + category."""
    results: list[dict] = await session_manager.get(req.session_id, _results_field(req.category))
    if results is None:
        raise HTTPException(status_code=404, detail="No audit results found. Run /audit-dynamic first.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Results"

    headers = [
        "Rule", "Status", "Criticality", "Confidence",
        "Pages", "Observation", "Recommendation", "Risk", "Requires Action",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    status_fills = {
        "Present":           PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "Partially Present": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
        "Inadequate":        PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
        "Not Present":       PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    }

    for r in results:
        row = [
            r.get("rule", ""),
            r.get("status", ""),
            r.get("criticality", ""),
            r.get("confidence_score", 0.0),
            r.get("page_numbers", ""),
            r.get("observation", ""),
            r.get("recommendation", ""),
            r.get("risk", ""),
            "Yes" if r.get("requires_action") else "No",
        ]
        ws.append(row)
        status = r.get("status", "")
        fill = status_fills.get(status)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    ws_sum = wb.create_sheet("Summary")
    summary_dict = verifier.get_result_summary(results)
    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["Total Rules", summary_dict["total_rules"]])
    ws_sum.append(["Compliance Rate (%)", summary_dict["compliance_rate"]])
    ws_sum.append(["Action Items", summary_dict["action_items"]])
    ws_sum.append(["High Priority Issues", summary_dict["high_priority_issues"]])
    ws_sum.append(["Average Confidence", summary_dict["average_confidence"]])
    for status, count in summary_dict["status_counts"].items():
        ws_sum.append([status, count])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"audit_{req.category.replace(' ', '_')}_{req.session_id[:8]}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# DELETE /clear/{session_id}
# ---------------------------------------------------------------------------

@router.delete("/clear/{session_id}")
async def clear_session(session_id: str):
    """Remove all audit data for a session (vector store + session keys)."""
    await session_manager.clear_session(session_id)
    await asyncio.to_thread(clear_collection, session_id)
    return {"status": "cleared", "session_id": session_id}
